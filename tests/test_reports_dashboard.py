from decimal import Decimal
from io import BytesIO
from time import monotonic
from django.test import TestCase,override_settings
from django.utils import timezone
from rest_framework.test import APIClient
from apps.accounts.models import User
from apps.inventory.posting import post_movement
from apps.locations.models import Location
from apps.master_data.models import Customer,FinishedProduct,ItemCategory,Supplier,UnitOfMeasurement
from apps.purchasing.models import PurchaseInvoice,SupplierLedger
from apps.sales.models import CustomerLedger,SalesInvoice,SalesInvoiceItem
from openpyxl import load_workbook

class ReportsDashboardTests(TestCase):
 def setUp(self):
  self.user=User.objects.create_user("report@test.local","password",full_name="Report",employee_code="REPORT",role="ADMINISTRATOR")
  self.client=APIClient();self.client.force_authenticate(self.user)
  self.unit=UnitOfMeasurement.objects.create(code="REP-U",name="Unit")
  category=ItemCategory.objects.create(name="REP-FG",kind="FG")
  self.product=FinishedProduct.objects.create(product_code="REP-P",name="Bread",category=category,sales_unit=self.unit,minimum_stock=Decimal("2"))
  self.location=Location.objects.create(code="REP-L",name="Shop",location_type="SHOP")
  self.customer=Customer.objects.create(customer_code="REP-C",name="Customer",opening_balance=Decimal("5"))
  self.supplier=Supplier.objects.create(supplier_code="REP-S",name="Supplier",opening_balance=Decimal("7"))
  post_movement(item=self.product,location=self.location,quantity=Decimal("10"),direction="IN",transaction_number="REP-IN",transaction_type="PRODUCTION_OUTPUT",reference_type="Test",reference_id=self.product.id,unit=self.unit,user=self.user,incoming_unit_cost=Decimal("2"))
  self.sale=SalesInvoice.objects.create(invoice_number="REP-SI",customer=self.customer,invoice_date=timezone.localdate(),sales_location=self.location,status="POSTED",grand_total=Decimal("50"),outstanding_amount=Decimal("30"),cost_of_goods_sold=Decimal("20"),gross_profit=Decimal("30"))
  SalesInvoiceItem.objects.create(sales_invoice=self.sale,finished_product=self.product,quantity=Decimal("5"),unit=self.unit,selling_price=Decimal("10"),line_total=Decimal("50"),cost_total=Decimal("20"),gross_profit=Decimal("30"))
  self.purchase=PurchaseInvoice.objects.create(invoice_number="REP-PI",supplier=self.supplier,invoice_date=timezone.localdate(),warehouse=self.location,status="POSTED",grand_total=Decimal("40"),outstanding_amount=Decimal("25"))
  CustomerLedger.objects.create(customer=self.customer,transaction_date=timezone.localdate(),reference_type="SalesInvoice",reference_id=self.sale.id,debit=Decimal("50"))
  SupplierLedger.objects.create(supplier=self.supplier,transaction_date=timezone.localdate(),reference_type="PurchaseInvoice",reference_id=self.purchase.id,credit=Decimal("40"))
 def test_inventory_and_ledger_reports_are_paginated_and_exportable(self):
  response=self.client.get("/api/reports/finished-goods-stock/?page_size=1");self.assertEqual(response.status_code,200);self.assertEqual(response.data["count"],1);self.assertEqual(response.data["results"][0]["quantity"],Decimal("10"))
  export=self.client.get("/api/reports/stock-ledger/?export=csv");self.assertEqual(export.status_code,200);self.assertEqual(export["Content-Type"],"text/csv")
 def test_financial_reports_reconcile_with_normalized_sources(self):
  customer=self.client.get("/api/reports/customer-outstanding/").data["results"][0];supplier=self.client.get("/api/reports/supplier-outstanding/").data["results"][0]
  self.assertEqual(customer["outstanding"],Decimal("55"));self.assertEqual(supplier["outstanding"],Decimal("47"))
 def test_dashboard_uses_backend_aggregates(self):
  response=self.client.get("/api/dashboard/");self.assertEqual(response.status_code,200);self.assertEqual(response.data["total_purchases"],Decimal("40"));self.assertEqual(response.data["daily_sales"],Decimal("50"));self.assertEqual(response.data["inventory_value"],Decimal("20"));self.assertEqual(response.data["receivables"],Decimal("30"));self.assertEqual(response.data["payables"],Decimal("25"));self.assertEqual(response.data["gross_profit"],Decimal("30"));self.assertEqual(response.data["top_products"][0]["sales"],Decimal("50"));self.assertEqual(response.data["top_customers"][0]["sales"],Decimal("50"))
 def test_dashboard_reconciles_with_purchase_sales_and_inventory_reports(self):
  dashboard=self.client.get("/api/dashboard/").data;purchase_rows=self.client.get("/api/reports/purchase-register/").data["results"];sales_rows=self.client.get("/api/reports/sales-register/").data["results"];inventory_rows=self.client.get("/api/reports/inventory-valuation/").data["results"]
  self.assertEqual(dashboard["total_purchases"],sum((row["grand_total"] for row in purchase_rows),Decimal("0")));self.assertEqual(dashboard["daily_sales"],sum((row["grand_total"] for row in sales_rows),Decimal("0")));self.assertEqual(dashboard["inventory_value"],sum((row["inventory_value"] for row in inventory_rows),Decimal("0")))
 def test_location_permission_rejects_another_location(self):
  restricted=User.objects.create_user("restricted-report@test.local","password",full_name="Restricted",employee_code="REPORT-R",role="MANAGER",assigned_location=self.location,allowed_modules=["reports"])
  other=Location.objects.create(code="REP-O",name="Other",location_type="SHOP");self.client.force_authenticate(restricted)
  response=self.client.get(f"/api/reports/finished-goods-stock/?location={other.id}");self.assertEqual(response.status_code,400)
 def test_excel_export_is_synchronous_and_does_not_create_a_job(self):
  response=self.client.post("/api/report-exports/",{"report_name":"sales-register","format":"XLSX","filters":{}},format="json")
  self.assertEqual(response.status_code,200);self.assertIn(".xlsx",response["Content-Disposition"])
  workbook=load_workbook(BytesIO(b"".join(response.streaming_content)),read_only=True);rows=list(workbook.active.values)
  self.assertEqual(rows[0][1],"invoice_number");self.assertEqual(rows[1][1],"REP-SI")
  from apps.reports.models import ReportExportJob
  self.assertFalse(ReportExportJob.objects.exists())
 def test_exports_neutralize_spreadsheet_formulas(self):
  self.product.name="=HYPERLINK(\"https://invalid\")";self.product.save(update_fields=["name"])
  response=self.client.post("/api/report-exports/",{"report_name":"sales-by-product","format":"XLSX","filters":{}},format="json")
  workbook=load_workbook(BytesIO(b"".join(response.streaming_content)),read_only=True);values=list(workbook.active.values)
  self.assertEqual(values[1][1],"'=HYPERLINK(\"https://invalid\")")
  csv_response=self.client.get("/api/reports/sales-by-product/?export=csv");content=b"".join(csv_response.streaming_content).decode()
  self.assertIn("'=HYPERLINK",content)
 @override_settings(REPORT_XLSX_MAX_ROWS=0)
 def test_excel_row_limit_directs_large_exports_to_streaming_csv(self):
  response=self.client.post("/api/report-exports/",{"report_name":"sales-register","format":"XLSX","filters":{}},format="json")
  self.assertEqual(response.status_code,400);self.assertIn("Use CSV",response.data["detail"])
  csv_response=self.client.post("/api/report-exports/",{"report_name":"sales-register","format":"CSV","filters":{}},format="json")
  self.assertEqual(csv_response.status_code,200);self.assertTrue(csv_response.streaming)
 @override_settings(REPORT_EXPORT_SPOOL_MAX_BYTES=1)
 def test_excel_export_spills_to_disk_and_completes_within_request_budget(self):
  from apps.reports.exports import build_xlsx
  started=monotonic();target=build_xlsx(({"name":f"row-{index}","value":index} for index in range(500)),max_rows=500)
  try:
   self.assertTrue(target._rolled);self.assertGreater(len(target.read()),0);self.assertLess(monotonic()-started,10)
  finally:target.close()
 def test_export_requires_authentication_and_enforces_location_scope(self):
  self.client.force_authenticate(user=None)
  self.assertEqual(self.client.post("/api/report-exports/",{"report_name":"sales-register","format":"XLSX"},format="json").status_code,401)
  other=Location.objects.create(code="REP-X",name="Other export",location_type="SHOP")
  restricted=User.objects.create_user("restricted-export@test.local","password",full_name="Restricted",employee_code="REPORT-X",role="MANAGER",assigned_location=self.location,allowed_modules=["reports"])
  self.client.force_authenticate(restricted)
  response=self.client.post("/api/report-exports/",{"report_name":"sales-register","format":"XLSX","filters":{"location":str(other.id)}},format="json")
  self.assertEqual(response.status_code,400)
